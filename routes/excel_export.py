from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from utils.data_loader import build_df_from_api
import openpyxl
import io
from openpyxl.utils import get_column_letter

router = APIRouter()

@router.get("/excel")
def get_excel():
    df = build_df_from_api()
    df = df.sort_values(by="id").reset_index(drop=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MKmax"

    headers = ["id", "level", "type", "title", "parent_id", "parent_name", "child_id", "body"]
    ws.append(headers)

    for i, row in df.iterrows():
        id_val = row["id"]
        type_val = "Attribute"
        title_val = row["title"]
        body_val = row["body"]
        row_idx = i + 2

        # Формулы с разделителем ";" для локали RU
        level_formula = f'=LEN(A{row_idx})-LEN(SUBSTITUTE(A{row_idx},".",""))+1'
        parent_id_formula = f'=IF(B{row_idx}=1,"",LEFT(A{row_idx},FIND("|",SUBSTITUTE(A{row_idx},".","|",B{row_idx}-1))-1))'
        parent_name_formula = f'=IF(C{row_idx}="", "", IFERROR(INDEX(D:D, MATCH(E{row_idx}, A:A, 0)), ""))'
        child_id_formula = (f'=TEXTJOIN(" | ", TRUE, FILTER(A:A, LEFT(A:A, LEN(A{row_idx})+1) = A{row_idx} & "."))'
        )

        ws.append([
            id_val,
            level_formula,
            type_val,
            title_val,
            parent_id_formula,
            parent_name_formula,
            child_id_formula,
            body_val
        ])

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 24

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=matrix_with_formulas.xlsx"
    })
