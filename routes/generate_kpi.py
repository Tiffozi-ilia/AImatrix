from fastapi import APIRouter, Query
from fastapi.responses import FileResponse
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from copy import copy

router = APIRouter()

@router.get("/generate_kpi")
def generate_kpi_report(
    year: int = Query(..., alias="year", description="Год, например 2025"),
    quarter: int = Query(..., ge=1, le=4, alias="quarter", description="Квартал: 1–4")
):
    quarters = {1: '1Q', 2: '2Q', 3: '3Q', 4: '4Q'}
    report_df = pd.read_excel("KPI.xlsx", skiprows=6)
    template_wb = openpyxl.load_workbook("EtalonKPI.xlsx")
    template_sheet = template_wb.active

    def is_full_quarter(date, year, quarter):
        if pd.isna(date):
            return False
        quarter_start = pd.Timestamp(year=year, month=(quarter - 1) * 3 + 1, day=1)
        return pd.Timestamp(date) < quarter_start

    def extract_template_style(sheet, row_idx=7):
        style_template = {}
        for col in range(2, 12):
            cell = sheet.cell(row=row_idx, column=col)
            style_template[col] = {
                "value": cell.value,
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "number_format": copy(cell.number_format),
                "protection": copy(cell.protection),
                "alignment": copy(cell.alignment)
            }
        return style_template

    def apply_template_style(sheet, target_row, template_style):
        for col, style in template_style.items():
            cell = sheet.cell(row=target_row, column=col)
            cell.value = style["value"]
            cell.font = style["font"]
            cell.border = style["border"]
            cell.fill = style["fill"]
            cell.number_format = style["number_format"]
            cell.protection = style["protection"]
            cell.alignment = style["alignment"]

    def sort_key_strict(index, row):
        indicator = str(row[report_df.columns[15]])
        division = str(row[report_df.columns[3]])
        position = str(row[report_df.columns[4]])
        perspective = str(row[report_df.columns[9]])

        if division == "Управление":
            if "NPS (ДРКИБ)" in indicator:
                return (0, index)
            elif "NPS (УБАИТ)" in indicator:
                return (1, index)
            elif "CSI (УБАИТ)" in indicator:
                return (2, index)
            elif perspective == "Повторяемая":
                return (3, index)
            else:
                return (4, index)
        elif position in ["Начальник", "Заместитель"]:
            if "NPS (УБАИТ)" in indicator:
                return (0, index)
            elif f"NPS ({division})" in indicator:
                return (1, index)
            elif f"CSI ({division})" in indicator:
                return (2, index)
            elif perspective == "Повторяемая":
                return (3, index)
            else:
                return (4, index)
        else:
            if "NPS (личный)" in indicator:
                return (0, index)
            elif f"NPS ({division})" in indicator:
                return (1, index)
            elif f"CSI ({division})" in indicator:
                return (2, index)
            elif perspective == "Повторяемая":
                return (3, index)
            else:
                return (4, index)

    positions_order = ["Начальник", "Заместитель", "руководитель направления",
                       "главный бизнес-аналитик", "ведущий бизнес-аналитик", "бизнес-аналитик"]

    filtered_df = report_df[(report_df.iloc[:, 2] == "Работает") &
                            (report_df.iloc[:, 6] == year) &
                            (report_df.iloc[:, 7] == quarters[quarter]) &
                            (report_df.iloc[:, 5].apply(lambda d: is_full_quarter(d, year, quarter)))]

    filtered_df["division_rank"] = filtered_df.iloc[:, 3].apply(lambda x: 0 if x == "Управление" else 1)
    filtered_df["position_rank"] = filtered_df.iloc[:, 4].apply(
        lambda x: positions_order.index(x) if x in positions_order else len(positions_order))
    df_final = filtered_df.sort_values(by=["division_rank", filtered_df.columns[3], "position_rank"])

    style_template = extract_template_style(template_sheet)

    for name in df_final.iloc[:, 1].unique():
        new_sheet = template_wb.copy_worksheet(template_sheet)
        new_sheet.title = name[:31]
        new_sheet["B1"] = f"Личная карточка индикаторов ({year})"
        new_sheet["B2"] = f"{quarter} квартал {year}"
        new_sheet["C3"] = name
        new_sheet["E3"] = f"KPI {quarter}Q{year}"

        df_person = df_final[df_final.iloc[:, 1] == name]
        rows_with_index = list(enumerate(df_person.to_dict('records')))
        sorted_rows = sorted(rows_with_index, key=lambda x: sort_key_strict(x[0], x[1]))
        start_row = 7

        for i, (_, kpi) in enumerate(sorted_rows):
            r = start_row + i
            new_sheet.insert_rows(r)
            apply_template_style(new_sheet, r, style_template)

            new_sheet[f"B{r}"] = kpi[report_df.columns[9]]
            new_sheet[f"C{r}"] = kpi[report_df.columns[14]]
            new_sheet[f"D{r}"] = kpi[report_df.columns[13]] / 100
            new_sheet[f"E{r}"] = kpi[report_df.columns[15]]
            new_sheet[f"F{r}"] = kpi[report_df.columns[16]]
            new_sheet[f"F{r}"].number_format = '0'
            new_sheet[f"G{r}"] = kpi[report_df.columns[17]]
            new_sheet[f"G{r}"].number_format = '0'
            new_sheet[f"H{r}"] = 1
            new_sheet[f"H{r}"].number_format = '0%'

            for col in ['I', 'J', 'K']:
                new_sheet[f"{col}{r}"] = template_sheet[f"{col}7"].value.replace('7', str(r))

            for col in range(2, 12):
                font9 = Font(name='Arial', size=9)
                new_sheet.cell(row=r, column=col).font = font9

        last_row = new_sheet.max_row
        last_data_row = max([r for r in range(1, last_row + 1) if new_sheet[f'E{r}'].value not in [None, '']])
        new_sheet.delete_rows(last_data_row + 1, last_row - last_data_row)

        total_row = last_data_row + 1
        new_sheet[f"I{total_row}"] = f"=SUM(I{start_row}:I{last_data_row})"
        new_sheet[f"I{total_row}"].number_format = '0%'
        new_sheet[f"K{total_row}"] = f"=SUM(K{start_row}:K{last_data_row})"
        new_sheet[f"K{total_row}"].number_format = '0%'

        sign_row = total_row + 2
        full_name = name.strip().split()
        short_name = f"{full_name[0]} {full_name[1][0]}." if len(full_name) >= 2 else name
        new_sheet[f"B{sign_row}"] = short_name
        new_sheet[f"F{sign_row}"] = "Панарин А."

    template_wb.remove(template_sheet)

    output_path = "Output_KPI_Report.xlsx"
    template_wb.save(output_path)

    return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="Output_KPI_Report.xlsx")
